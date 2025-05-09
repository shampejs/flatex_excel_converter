from openpyxl import Workbook
from openpyxl.styles import PatternFill

from trade_class import Trade
from trading_action_class import TradingAction
import copy

colors = ['ADD8E6', '90EE90', 'FFFACD', 'FFB6C1', 'E6E6FA', 'F5F5DC']


def create_excel_file(transactions, name_map):
    wb = Workbook()  # Create a workbook

    create_historic_sheet(wb, transactions)
    create_sorted_sheet(wb, transactions, name_map)
    create_detailed_overview(wb, transactions, name_map)

    wb.save("generated_flatex.xlsx")

def create_historic_sheet(wb, transactions):
    ws = wb.active  # Get the active sheet
    ws.title = "Historisch"  # Rename sheet
    ws.append(
        ["Datum", "Name", "Typ", "Kurs", "Anzahl", "Gesamt", "Provision", "Eigene Spesen", "Fremnde Spesen", "KEST",
         "Endbetrag", "Gewinn + / Verlust -"])

    for transaction in transactions:
        if transaction:
            row_data = transaction.get_excel_format()
            ws.append(row_data)

            # Get the row number of the newly appended row
            row_number = ws.max_row

            # Color the entire row
            color_row(ws, row_number, colors[get_month_of_date(transaction.date) % len(colors)])

def create_sorted_sheet(wb, transactions, name_map):
    ws = wb.create_sheet("Sortiert")
    ws.append(
        ["Datum", "Name", "Typ", "Kurs", "Anzahl", "Gesamt", "Provision", "Eigene Spesen", "Fremnde Spesen", "KEST",
         "Endbetrag", "Gewinn + / Verlust -"])

    last_name = transactions[0].name

    for transaction in sorted(transactions, key=lambda x: x.name):
        if last_name != transaction.name:
            ws.append([])

        if transaction:
            ws.append(transaction.get_excel_format())
            # Get the row number of the newly appended row
            row_number = ws.max_row

            # Color the entire row
            color_row(ws, row_number, colors[name_map[transaction.name] % len(colors)])

        last_name = transaction.name





def create_detailed_overview(wb, transactions, name_map):
    ws = wb.create_sheet("Detailiert")
    ws.append(
        ["Datum", "Name", "Typ", "Kurs", "Anzahl", "Gesamt", "Provision", "Eigene Spesen", "Fremnde Spesen", "KEST",
         "Endbetrag", "Gewinn + / Verlust -"])

    tmp_transaction = None
    last_name = transactions[0].name

    for transaction in sorted(transactions, key=lambda x: x.name):
        if last_name != transaction.name:
            ws.append([])
            tmp_transaction = None

        if transaction:
            if tmp_transaction:
                # calculate the pure win / loss
                if transaction.kind_of_transaction == "Verkauf":
                    create_billing(wb, transaction, tmp_transaction, name_map)

                tmp_transaction.update_transaction(transaction)

            else:
                tmp_transaction = TradingAction(transaction)

            add_line_to_excel(ws,transaction.get_excel_format(), colors[name_map[transaction.name] % len(colors)])
            add_line_to_excel(ws,tmp_transaction.get_excel_format(), colors[name_map[transaction.name] % len(colors)])

            # Update the tmp_transaction to take action on the last transaction

        last_name = transaction.name

def create_billing(wb, transaction, tmp_transaction, name_map):
    if "Abrechnung" in wb.sheetnames:
        ws = wb.get_sheet_by_name("Abrechnung")
    else:
        ws = wb.create_sheet("Abrechnung")
        ws.append(
            ["Datum", "Name", "Typ", "Kurs", "Anzahl", "Gesamt", "Provision", "Eigene Spesen", "Fremnde Spesen", "KEST",
             "Endbetrag", "Gewinn + / Verlust -"])
    add_line_to_excel(ws, tmp_transaction.get_excel_format(), colors[name_map[transaction.name] % len(colors)])
    add_line_to_excel(ws, transaction.get_excel_format(), colors[name_map[transaction.name] % len(colors)])


    # calculate win or loss
    selling_transaction = copy.deepcopy(tmp_transaction)
    selling_transaction.pre_selling_calculation(transaction)
    add_line_to_excel(ws, selling_transaction.get_excel_format(), colors[name_map[transaction.name] % len(colors)])
    selling_transaction.selling_result(transaction)
    add_line_to_excel(ws, selling_transaction.get_excel_format(), colors[name_map[transaction.name] % len(colors)])

    ws.append([])

def add_line_to_excel(ws, line, color):
    ws.append(line)
    color_row(ws, ws.max_row, color)


def color_row(worksheet, row_index, color_hex):
    """Colors all cells in a given row with the specified hex color."""
    fill = PatternFill(fill_type='solid', start_color=color_hex)
    for cell in worksheet[row_index]:
        cell.fill = fill

def get_month_of_date(date):
    return int(date.split(".")[1])